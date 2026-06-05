"""Extrator GENÉRICO de permutador casco-tubo (BEU, BEM, ...) da planilha ENGEMATEX.

Uso:  python extract_permutador.py BEU
      python extract_permutador.py BEM

Produz em pricing_engine/seeds/:
  {d}_ground_truth.json  — todos os blocos + totais (gabarito de custo)
  {d}_materiais.json     — blocos das seções de matéria-prima (geometrizáveis + catálogo)
  {d}_operacoes.json     — blocos de fabricação/finalização (mão-de-obra vs serviço)

Anatomia (idêntica entre BEU/BEM): aba 'PERMUTADOR "BEM" - ORÇAMENTO', col 84 = PREÇO R$,
col 72 = peso líq, col 78 = peso bruto. Cada bloco = header (rótulos) + linha de dados.
Seções carregadas dos cabeçalhos: 'MATÉRIA PRIMA - {FEIXE,CASCO,CABEÇOTE}',
'FABRICAÇÃO...', 'ENGENHARIA...', 'FERRAMENTAS...'.
"""
import glob, json, os, sys
import openpyxl

SEEDS = os.path.join(os.path.dirname(__file__), "..", "pricing_engine", "seeds")
GABARITOS = {  # totais lidos da planilha (custo c/ imp, venda c/ imp, venda s/ imp, F.C., ICMS)
    "BEU": dict(custo=128160.0, venda_com=160200.0, venda_sem=146103.0, fc=1.25, icms=9.0,
                peso_liq=3016.0, peso_bruto=None, desc="BEU (bonnet + casco 1 passe + feixe em U)"),
    "BEM": dict(custo=119295.0, venda_com=149119.0, venda_sem=135997.0, fc=1.25, icms=9.0,
                peso_liq=3021.0, peso_bruto=3373.0, desc="BEM (bonnet + casco 1 passe + cabeçote traseiro fixo, tubos retos)"),
}


def secao_de(texto):
    if not isinstance(texto, str):
        return None
    t = texto.upper()
    if "MATÉRIA PRIMA" in t and "FEIXE" in t:
        return "feixe_material"
    if "MATÉRIA PRIMA" in t and "CASCO" in t:
        return "casco_material"
    if "MATÉRIA PRIMA" in t and "CABE" in t:
        return "cabecote_material"
    if t.startswith("FABRICAÇÃO"):
        return "fabricacao"
    if t.startswith("ENGENHARIA") or t.startswith("FERRAMENTAS"):
        return "finalizacao"
    return None


def familia(label, dims):
    """Classifica o material numa família de PESO. 'perfurado' = espelho/chicana (placa
    com furos ou segmento recortado) — peso vem do seed (não é disco cheio); geometria
    desses é do motor do feixe. 'catalogo' = item de preço fixo (comercial)."""
    L = (label or "").upper()
    keys = set(dims)
    if "TUBO" in L and "TROCA" in L:
        return "tubo"
    if "TAMPO" in L:
        return "tampo_2_1"
    if "PESCOÇO" in L or "PESCOCO" in L:
        return "pipe"
    if "FLANGE" in L and "PRINCIPAL" in L:
        return "anel"
    if "FLANGE" in L:
        return "flange_wn"
    if "ANEL" in L:
        return "anel"
    if "VIROLA" in L:
        return "chapa_retangular"
    if "SUPORTE" in L:
        return "catalogo"
    # espelho (furado) e chicana/segmento (TRANSVERSAL) não são discos cheios
    if "ESPELHO" in L or "CHICANA" in L or "TRANSVERSAL" in L:
        return "perfurado"
    if "DIVISORA" in L or "CHAPA" in L:
        return "chapa_retangular"
    if {"OD", "ID", "ESP."} <= keys:
        return "anel"
    if {"ESP.", "LARGURA", "COMPR."} <= keys or {"ESP.", "LARG.", "COMPR."} <= keys:
        return "chapa_retangular"
    if {"OD", "ESP.", "COMPR."} <= keys:
        return "tubo"
    if {"OD", "ESP."} <= keys:
        return "disco"
    return "catalogo"


def _aba_orcamento(wb, designacao):
    """Acha a aba de ORÇAMENTO de forma robusta (não depende do literal 'BEM').
    Preferência: nome contendo a designação + ORÇAMENTO; senão a única com ORÇAMENTO;
    senão o nome legado. NOTA: os offsets i-3/i-2 do parser dependem do layout Rev.0."""
    nomes = wb.sheetnames
    up = designacao.upper()
    for n in nomes:
        u = n.upper()
        if "ORÇAMENTO" in u and up in u:
            return wb[n]
    orc = [n for n in nomes if "ORÇAMENTO" in n.upper()]
    if len(orc) == 1:
        return wb[orc[0]]
    if 'PERMUTADOR "BEM" - ORÇAMENTO' in nomes:
        return wb['PERMUTADOR "BEM" - ORÇAMENTO']
    if orc:
        return wb[orc[0]]
    raise KeyError(f"Aba de ORÇAMENTO não encontrada em {nomes}")


def extrair(designacao):
    p = glob.glob(f"/home/rcosta00/dev/uploads/*{designacao}*.xlsx")[0]
    wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
    ws = _aba_orcamento(wb, designacao)
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=90, values_only=True))

    def cell(r, c):
        return rows[r - 1][c - 1] if 1 <= r <= len(rows) else None

    DIM_SKIP = {"P.E.", "PESO\nESPECÍF.", "PESO LÍQ.", "PESO BR.", "PREÇO R$", "PESO LIQ."}
    ground, materiais, operacoes = [], [], []
    secao = None
    code_seen = {}
    for i in range(1, len(rows) + 1):
        # atualiza seção corrente a partir de cabeçalhos em col2/col7
        for c in (2, 7):
            s = secao_de(cell(i, c))
            if s:
                secao = s
        preco = cell(i, 84)
        if not isinstance(preco, (int, float)) or preco == 0:
            continue
        # rótulo
        label = None
        for rr, c in ((i, 7), (i, 2), (i - 2, 7), (i - 2, 2)):
            v = cell(rr, c)
            if isinstance(v, str) and v.strip() and v.strip() not in ("x", "SIM", "NÃO", "APLICÁVEL"):
                label = v.strip().replace("\n", " "); break
        peso_br = cell(i, 78)
        peso_liq = cell(i, 72)
        preco = round(float(preco), 2)
        ground.append({"row": i, "secao": secao, "label": label, "preco": preco})

        if secao and secao.endswith("_material"):
            # header de dimensões 2 linhas acima
            hdr = {}
            for ci, v in enumerate(rows[i - 3], start=1):
                if isinstance(v, str) and v.strip() and v.strip() not in ("x",):
                    hdr[v.strip()] = ci
            dims = {}
            for name, c in hdr.items():
                if name in DIM_SKIP:
                    continue
                val = cell(i, c)
                if isinstance(val, (int, float, str)):
                    dims[name] = round(val, 3) if isinstance(val, float) else val
            material = None
            for c in (54,):
                v = cell(i - 2, c) or cell(i, c)
                if isinstance(v, str) and v.strip():
                    material = v.strip()
            fam = familia(label, dims)
            m = {"row": i, "secao": secao, "label": label, "familia": fam,
                 "material": material, "dims": dims, "preco": preco}
            if isinstance(peso_br, (int, float)) and peso_br:
                m["peso_bruto"] = round(float(peso_br), 3)
                m["peso_liq"] = round(float(peso_liq), 3) if isinstance(peso_liq, (int, float)) else None
                m["price_kgf"] = round(preco / float(peso_br), 3)
            else:
                m["familia"] = "catalogo"   # item de preço fixo (placa, etc.)
                m["peso_bruto"] = None
            materiais.append(m)
        else:
            # operação: decompõe horas×rate lendo o header
            hdr = {}
            for ci, v in enumerate(rows[i - 3], start=1):
                if isinstance(v, str) and v.strip() and v.strip() not in ("x",):
                    hdr[v.strip()] = ci
            horas = cell(i, hdr["HORAS"]) if "HORAS" in hdr else None
            rate = cell(i, hdr["R$ / HORA"]) if "R$ / HORA" in hdr else None
            ajuste = (cell(i, hdr["AJUSTE"]) if "AJUSTE" in hdr else 0) or 0
            is_labor = isinstance(horas, (int, float)) and isinstance(rate, (int, float))
            base = "".join(ch for ch in (label or "OP").upper() if ch.isalnum())[:18]
            code_seen[base] = code_seen.get(base, 0) + 1
            o = {"code": f"{(secao or 'OP')[:3].upper()}-{base}-{code_seen[base]}",
                 "row": i, "secao": secao, "label": label,
                 "tipo": "mao_obra" if is_labor else "servico",
                 "preco_gabarito": preco, "ajuste": round(float(ajuste), 2)}
            if is_labor:
                o["horas"] = round(float(horas), 3)
                o["rate"] = round(float(rate), 2)
            operacoes.append(o)

    return p, ground, materiais, operacoes


def main():
    d = (sys.argv[1] if len(sys.argv) > 1 else "BEU").upper()
    g = GABARITOS[d]
    src, ground, materiais, operacoes = extrair(d)
    soma = round(sum(b["preco"] for b in ground), 2)
    por_secao = {}
    for b in ground:
        por_secao[b["secao"]] = round(por_secao.get(b["secao"], 0) + b["preco"], 2)

    doc_g = {"fonte": os.path.basename(src), "designacao_tema": d, "descricao": g["desc"],
             "custo_total_com_impostos": g["custo"], "preco_venda_com_impostos": g["venda_com"],
             "preco_venda_sem_impostos": g["venda_sem"], "fator_comercial": g["fc"],
             "icms_pct": g["icms"], "peso_liquido_total_kgf": g["peso_liq"],
             "soma_blocos": soma, "n_blocos": len(ground), "por_secao": por_secao, "blocos": ground}
    labor = [o for o in operacoes if o["tipo"] == "mao_obra"]
    doc_o = {"fonte": os.path.basename(src), "designacao_tema": d, "n_ops": len(operacoes),
             "n_mao_obra": len(labor), "n_servico": len(operacoes) - len(labor),
             "operacoes": operacoes}
    doc_m = {"fonte": os.path.basename(src), "designacao_tema": d,
             "n_materiais": len(materiais), "materiais": materiais}

    for nome, doc in ((f"{d.lower()}_ground_truth.json", doc_g),
                      (f"{d.lower()}_operacoes.json", doc_o),
                      (f"{d.lower()}_materiais.json", doc_m)):
        with open(os.path.join(SEEDS, nome), "w", encoding="utf-8") as f:
            json.dump(doc, f, ensure_ascii=False, indent=2)

    print(f"[{d}] {os.path.basename(src)}")
    print(f"  blocos={len(ground)}  soma=R$ {soma:,.2f}  gabarito=R$ {g['custo']:,.2f}  Δ={(soma-g['custo'])/g['custo']:+.2%}")
    print(f"  materiais={len(materiais)}  operacoes={len(operacoes)} ({len(labor)} MO / {len(operacoes)-len(labor)} serviço)")
    for s, v in por_secao.items():
        print(f"    {s:20} R$ {v:>12,.2f}")


if __name__ == "__main__":
    main()
