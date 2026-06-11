"""Extrai a tabela de peso de flanges Welding Neck (WN) da aba 'FLANGES WN' da planilha
ENGEMATEX → pricing_engine/seeds/flanges_wn.json.

Peso (kgf/peça, aço-carbono) por CLASSE de pressão × NPS × SCHEDULE. Resposta do Wellington
(A3): o motor deve puxar o peso real do flange por Ø × rating, em vez de chutar — senão erra
o peso final e as horas de solda dos bocais. Fonte: blocos "PESO DE FLANGES WELDING NECK 'WN'
X#" (face RF). Confirmado exato vs gabarito: 8\" SCH80 600# = 56 kg; 10\" SCH40 600# = 86,8 kg.
"""
import glob, json, os, re

import openpyxl

SRC = glob.glob("/home/rcosta00/dev/uploads/*BEU*.xlsx")[0]
OUT = os.path.join(os.path.dirname(__file__), "..", "pricing_engine", "seeds", "flanges_wn.json")


def _norm_sched(s):
    """Normaliza rótulo de schedule (5S, 10, 10S, 20, STD, 40, XS, 80, 80S...)."""
    return str(s).strip().upper().replace(".0", "")


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True, read_only=True)
    ws = wb["FLANGES WN"]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=20, values_only=True))

    tabela = {}   # {classe: {nps: {sched: kgf}}}
    i = 0
    while i < len(rows):
        c2 = rows[i][1]
        if isinstance(c2, str) and "WELDING NECK" in c2.upper() and "RF" in c2.upper():
            m = re.search(r"(\d+)#", c2)
            if not m:
                i += 1
                continue
            classe = f"{m.group(1)}#"
            # header de schedules = próxima linha cujo col2 == 'NPS'
            h = i + 1
            while h < len(rows) and str(rows[h][1]).strip().upper() != "NPS":
                h += 1
            if h >= len(rows):
                break
            header = rows[h]
            # mapa coluna(0-based) → schedule, a partir da col E (idx4) até onde houver rótulo
            sched_cols = {}
            for ci in range(4, 18):
                lbl = header[ci]
                if lbl not in (None, "", "OD"):
                    sched_cols[ci] = _norm_sched(lbl)
            # linhas de dados: NPS na col2 (ex '8"'), até a próxima linha vazia/sem NPS válido
            tabela.setdefault(classe, {})
            r = h + 1
            while r < len(rows):
                nps = rows[r][1]
                # NPS é curto e termina em " (ex 8", 1/2", 1.1/4"). O título do próximo bloco
                # também tem aspas ("WN") mas é longo → para o bloco nele.
                is_nps = isinstance(nps, str) and nps.strip().endswith('"') and len(nps.strip()) <= 8
                if not is_nps:
                    if isinstance(nps, str) and "WELDING NECK" in nps.upper():
                        break
                    r += 1
                    if r - h > 40:
                        break
                    continue
                nps = nps.strip()
                pesos = {}
                for ci, sched in sched_cols.items():
                    v = rows[r][ci]
                    if isinstance(v, (int, float)) and v > 0:
                        pesos[sched] = round(float(v), 1)
                if pesos:
                    tabela[classe][nps] = pesos
                r += 1
            i = r
        else:
            i += 1

    doc = {"fonte": os.path.basename(SRC), "tipo": "WN RF (Welding Neck Raised Face)",
           "unidade": "kgf/peça", "material_base": "aço-carbono (corrigir densidade p/ liga)",
           "classes": sorted(tabela.keys()), "tabela": tabela}
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(doc, f, ensure_ascii=False, indent=2)

    print(f"Gravado {OUT}")
    print(f"  classes: {sorted(tabela.keys())}")
    for cl in sorted(tabela):
        print(f"    {cl}: {len(tabela[cl])} NPS")
    # checagens contra o gabarito
    def look(cl, nps, sch):
        return tabela.get(cl, {}).get(nps, {}).get(sch)
    c8 = look('600#', '8"', '80')
    c10 = look('600#', '10"', '40')
    c6 = look('600#', '6"', 'STD')
    print("  check 8\" 600# SCH80 =", c8, "(esperado ~56)")
    print("  check 10\" 600# SCH40 =", c10, "(esperado ~86,8)")
    print("  check 6\" 600# STD =", c6)


if __name__ == "__main__":
    main()
